#!/usr/bin/env python3
"""
build_reference_assets.py — Écluse, Jalon 2
Transforme les fichiers publics bruts (INSEE COG, populations communales,
pyramide des ages, DREES demographie, prenoms INSEE) en assets compacts pour
le package ecluse_reference. Deterministe et rejouable.

Entrees attendues dans RAW_DIR :
  - v_commune_2026__1_.csv                      (COG communes)
  - ensemble__1_.zip -> donnees_communes.csv    (populations communales)
  - Population-2025.zip -> 3_Pop1janv_age.xlsx  (pyramide des ages)
  - Medecins_RPPS_2012-2026.xlsx                (specialite x departement)
  - {Pharmaciens,Chirurgiens-dentistes,Sages-femmes,Pedicures-podologues}_RPPS_*.xlsx
  - prenoms-2025-liste_csv.zip -> prenoms-2025-liste.csv
"""
import csv, gzip, json, os, sys, unicodedata, zipfile
from collections import defaultdict
from openpyxl import load_workbook

RAW = "/mnt/user-data/uploads"
OUT = "/home/claude/assets"
TMP = "/home/claude/reftmp"
os.makedirs(OUT, exist_ok=True)
os.makedirs(TMP, exist_ok=True)

def norm_name(s: str) -> str:
    if s is None: return ""
    s = unicodedata.normalize("NFKD", str(s)).encode("ascii", "ignore").decode()
    s = s.upper()
    for ch in "'-\u2019./":
        s = s.replace(ch, " ")
    return " ".join(s.split())

def norm_dep(raw: str):
    """DREES '057 -Moselle' / '02A -Corse-du-Sud' / '971 -Guadeloupe' -> '57'/'2A'/'971'."""
    if raw is None: return None
    code = str(raw).split("-", 1)[0].strip()
    if code in ("000", "999"): return None
    if len(code) == 3 and code[0] == "0":
        code = code[1:]
    return code

def unzip_one(zip_name, inner):
    path = os.path.join(RAW, zip_name)
    with zipfile.ZipFile(path) as z:
        z.extract(inner, TMP)
    return os.path.join(TMP, inner)

ENS0, ENS00 = "0-Ensemble", "00-Ensemble"
report = {}

# ---------- A + B : geo communes + populations ----------
print("== Geo + populations ==")
communes = {}  # COM -> [LIBELLE, DEP, REG]
with open(os.path.join(RAW, "v_commune_2026__1_.csv"), encoding="utf-8") as f:
    for row in csv.DictReader(f):
        if row["TYPECOM"] == "COM":
            communes[row["COM"]] = [row["LIBELLE"], row["DEP"], row["REG"]]
print("  communes (TYPECOM=COM):", len(communes))

pop = {}                # COM -> PMUN
reg_label = {}          # REG -> libelle
pop_csv = unzip_one("ensemble__1_.zip", "donnees_communes.csv")
with open(pop_csv, encoding="utf-8") as f:
    for row in csv.DictReader(f, delimiter=";"):
        pop[row["COM"]] = int(row["PMUN"])
        reg_label[row["REG"]] = row["Région"]
missing_pop = [c for c in communes if c not in pop]
print("  communes sans PMUN:", len(missing_pop))

# dep -> reg
dep_reg = {}
for com, (lib, dep, reg) in communes.items():
    dep_reg.setdefault(dep, reg)

# ecrire geo_communes.csv.gz  (COM;DEP;REG;PMUN;NORM;LIBELLE) trie par COM
geo_path = os.path.join(OUT, "geo_communes.csv.gz")
with gzip.open(geo_path, "wt", encoding="utf-8", newline="") as gz:
    w = csv.writer(gz, delimiter=";")
    w.writerow(["COM", "DEP", "REG", "PMUN", "NORM", "LIBELLE"])
    for com in sorted(communes):
        lib, dep, reg = communes[com]
        w.writerow([com, dep, reg, pop.get(com, ""), norm_name(lib), lib])
report["geo_communes"] = len(communes)

with open(os.path.join(OUT, "geo_hierarchie.json"), "w", encoding="utf-8") as f:
    json.dump({"dep_reg": dep_reg, "reg_label": reg_label}, f, ensure_ascii=False, indent=0)

# ---------- C : medecins (specialite x departement) ----------
print("== Medecins ==")
def load_effectifs(path, cross_col):
    """Retourne {(dep, code_cross): (label_cross, effectif_2026)} avec autres dims=Ensemble."""
    wb = load_workbook(path, read_only=True)
    ws = wb["Effectifs"]
    it = ws.iter_rows(values_only=True)
    h = list(next(it))
    idx = {c: i for i, c in enumerate(h) if c}
    i_dep, i_cross = idx["departement"], idx[cross_col]
    i_exo, i_age, i_sexe = idx["exercice"], idx["tranche_age"], idx["sexe"]
    i_2026 = idx["effectif_2026"]
    out = {}
    for r in it:
        if not (str(r[i_exo]) == ENS0 and str(r[i_age]) == ENS00 and str(r[i_sexe]) == ENS0):
            continue
        dep = norm_dep(r[i_dep])
        cross = r[i_cross]
        if dep is None or cross is None or str(cross) == ENS00:
            continue
        code, _, label = str(cross).partition("-")
        val = r[i_2026]
        if val is None: continue
        out[(dep, code.strip())] = (label.strip(), int(val))
    wb.close()
    return out

med = load_effectifs(os.path.join(RAW, "Médecins_RPPS_2012-2026.xlsx"), "specialites")
rar_med = defaultdict(dict); spec_label = {}
for (dep, code), (label, val) in med.items():
    rar_med[dep][code] = val
    spec_label[code] = label
with open(os.path.join(OUT, "rarity_medecins.json"), "w", encoding="utf-8") as f:
    json.dump({"labels": spec_label, "by_dep": rar_med}, f, ensure_ascii=False)
print("  paires (dep x specialite):", len(med), "| specialites:", len(spec_label))
report["rarity_medecins_pairs"] = len(med)

# ---------- D : autres professions RPPS (profession x departement) ----------
print("== Autres professions RPPS ==")
PROF_FILES = {
    "pharmacien": "Pharmaciens_RPPS_2012-2026.xlsx",
    "chirurgien-dentiste": "Chirurgiens-dentistes_RPPS_2012-2026.xlsx",
    "sage-femme": "Sages-femmes_RPPS_2012-2026.xlsx",
    "pedicure-podologue": "Pédicures-podologues_RPPS_2018-2026.xlsx",
}
rar_prof = defaultdict(dict)  # profession -> {dep: effectif}
for prof, fn in PROF_FILES.items():
    path = os.path.join(RAW, fn)
    wb = load_workbook(path, read_only=True); ws = wb["Effectifs"]
    it = ws.iter_rows(values_only=True); h = list(next(it))
    idx = {c: i for i, c in enumerate(h) if c}
    i_dep = idx["departement"]; i_sect = idx["secteur_activite"]
    i_exo, i_age, i_sexe = idx["exercice"], idx["tranche_age"], idx["sexe"]
    i_2026 = idx["effectif_2026"]
    n = 0
    for r in it:
        if not (str(r[i_sect]) == ENS00 and str(r[i_exo]) == ENS0
                and str(r[i_age]) == ENS00 and str(r[i_sexe]) == ENS0):
            continue
        dep = norm_dep(r[i_dep])
        if dep is None: continue
        val = r[i_2026]
        if val is None: continue
        rar_prof[prof][dep] = int(val); n += 1
    wb.close()
    print(f"  {prof:22s}: {n} departements")
with open(os.path.join(OUT, "rarity_professions.json"), "w", encoding="utf-8") as f:
    json.dump(rar_prof, f, ensure_ascii=False)
report["rarity_professions"] = {k: len(v) for k, v in rar_prof.items()}

# ---------- E : pyramide des ages (national 2026) ----------
print("== Pyramide des ages ==")
age_path = unzip_one("Population-2025.zip", "3_Pop1janv_age.xlsx")
wb = load_workbook(age_path, read_only=True)
ws = wb["2026"]
rows = list(ws.iter_rows(values_only=True))
# ligne d'entete = celle qui contient la cellule "Année de naissance"
hdr = next(i for i, r in enumerate(rows)
           if any(c and str(c).strip() == "Année de naissance" for c in r))
cols = [str(c).strip() if c else "" for c in rows[hdr]]
i_age = next(i for i, c in enumerate(cols) if "révolues" in c)
# deux colonnes "Ensemble" : metropole puis France entiere -> prendre la DERNIERE (France entiere)
i_ens = max(i for i, c in enumerate(cols) if c == "Ensemble")
age_pyr = {}
for r in rows[hdr+1:]:
    a = r[i_age]
    if a is None: continue
    try: age = int(str(a).strip())
    except ValueError: continue
    v = r[i_ens]
    if v is None: continue
    age_pyr[age] = int(v)
wb.close()
with open(os.path.join(OUT, "age_pyramide.json"), "w", encoding="utf-8") as f:
    json.dump(age_pyr, f)
print("  ages:", len(age_pyr), "| total metropole:", sum(age_pyr.values()))
report["age_pyramide_total"] = sum(age_pyr.values())

# ---------- F : gazetteer prenoms ----------
print("== Prenoms ==")
pren_csv = unzip_one("prenoms-2025-liste_csv.zip", "prenoms-2025-liste.csv")
prenoms = set()
with open(pren_csv, encoding="utf-8") as f:
    for row in csv.DictReader(f, delimiter=";"):
        p = norm_name(row["prenom"])
        if len(p) >= 3:
            prenoms.add(p)
with gzip.open(os.path.join(OUT, "first_names.txt.gz"), "wt", encoding="utf-8") as gz:
    for p in sorted(prenoms):
        gz.write(p + "\n")
print("  prenoms normalises (>=3 car.):", len(prenoms))
report["first_names"] = len(prenoms)

# ---------- meta ----------
with open(os.path.join(OUT, "reference_meta.json"), "w", encoding="utf-8") as f:
    json.dump({
        "millesimes": {"cog": 2026, "populations": "ref2023->2026", "drees": 2026,
                       "pyramide": "2026 (France entiere)", "prenoms": 2025},
        "approximations": ["pyramide France entiere 2026", "specialites medecins niveau medium",
                           "prenoms sans frequence (signal de confiance, pas masque autonome)"],
        "report": report,
    }, f, ensure_ascii=False, indent=2)

# ---------- ASSERTIONS DE CONTROLE (le filet) ----------
print("\n== Assertions de controle ==")
assert len(communes) == 34875, f"communes attendu 34875, obtenu {len(communes)}"
assert rar_med["57"]["11"] == 51, f"Pneumologie/Moselle attendu 51, obtenu {rar_med['57'].get('11')}"
orphans = [(d, c) for (d, c) in med if d not in dep_reg]
assert not orphans, f"departements DREES orphelins (hors COG): {orphans[:5]}"
print("  OK communes == 34875")
print("  OK Pneumologie/Moselle (57,11) ==", rar_med["57"]["11"])
print("  OK aucun departement DREES orphelin")
print("\nTermine. Assets dans", OUT)
